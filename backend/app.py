from flask import Flask, request, jsonify, render_template, send_file, after_this_request
import boto3
from datetime import datetime, date
import openpyxl
import os
import json
from io import BytesIO
from PIL import Image
import shutil

app = Flask(__name__)

# ========= Paths (absolute) =========
BASE_DIR = app.root_path
DATA_DIR = os.path.join(BASE_DIR, "data")
REF_DIR = os.path.join(BASE_DIR, "reference_images")
os.makedirs(DATA_DIR, exist_ok=True)
os.makedirs(REF_DIR, exist_ok=True)

ATTENDANCE_FILE = os.path.join(DATA_DIR, "attendance.xlsx")
STUDENT_REGISTRY = os.path.join(DATA_DIR, "students.json")

# ========= AWS Rekognition =========
rekognition = boto3.client("rekognition", region_name="us-east-1")
COLLECTION_ID = "students_collection"

# ========= Session memory =========
POLL_INTERVAL_SEC = 2  # must match frontend polling
SESSION_PRESENT_SET = set()     # { "student_id:date" }
FOCUS_SECONDS = {}              # student_id -> seconds
DISTRACTED_SECONDS = {}         # student_id -> seconds


# ========= Init helpers =========
def ensure_collection():
    try:
        rekognition.describe_collection(CollectionId=COLLECTION_ID)
    except rekognition.exceptions.ResourceNotFoundException:
        rekognition.create_collection(CollectionId=COLLECTION_ID)


def _safe_save_wb(wb, path):
    """
    Save workbook safely even if Excel has the file open.
    Writes to *_autosave.xlsx first, then copies.
    """
    try:
        wb.save(path)
    except PermissionError:
        tmp = path.replace(".xlsx", "_autosave.xlsx")
        wb.save(tmp)
        try:
            shutil.copyfile(tmp, path)
        except Exception:
            pass


def ensure_attendance_file():
    """Create attendance.xlsx if not exists."""
    if not os.path.exists(ATTENDANCE_FILE) or os.path.getsize(ATTENDANCE_FILE) == 0:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Attendance"
        ws.append(["StudentID", "Name", "Date", "Time", "Status", "Concentration"])
        _safe_save_wb(wb, ATTENDANCE_FILE)


def ensure_student_registry():
    """Create students.json if not exists."""
    if not os.path.exists(STUDENT_REGISTRY):
        with open(STUDENT_REGISTRY, "w") as f:
            json.dump({}, f, indent=2)


ensure_collection()
ensure_attendance_file()
ensure_student_registry()


# ========= Utility functions =========
def get_student_name(student_id: str) -> str:
    with open(STUDENT_REGISTRY, "r") as f:
        registry = json.load(f)
    return registry.get(student_id, "Unknown")


def save_student(student_id: str, name: str):
    with open(STUDENT_REGISTRY, "r") as f:
        registry = json.load(f)
    registry[student_id] = name
    with open(STUDENT_REGISTRY, "w") as f:
        json.dump(registry, f, indent=2)


def already_logged_today(student_id: str) -> bool:
    """Check if student already marked present today."""
    today_key = f"{student_id}:{date.today().isoformat()}"
    if today_key in SESSION_PRESENT_SET:
        return True

    wb = openpyxl.load_workbook(ATTENDANCE_FILE)
    ws = wb["Attendance"]
    for row in ws.iter_rows(min_row=2, values_only=True):
        sid, _, d, *_ = row
        if sid == student_id and d == date.today().isoformat():
            SESSION_PRESENT_SET.add(today_key)
            return True
    return False


def log_attendance_once(student_id: str, name: str, concentration_label: str):
    """Log one row per student/day (skip 'unknown')."""
    if student_id == "unknown":
        return
    if already_logged_today(student_id):
        return

    wb = openpyxl.load_workbook(ATTENDANCE_FILE)
    ws = wb["Attendance"]

    now = datetime.now()
    ws.append([
        student_id,
        name,
        now.strftime("%Y-%m-%d"),
        now.strftime("%H:%M:%S"),
        "Present",
        concentration_label
    ])
    _safe_save_wb(wb, ATTENDANCE_FILE)

    SESSION_PRESENT_SET.add(f"{student_id}:{date.today().isoformat()}")


def pil_crop_from_normalized_bbox(img: Image.Image, bbox: dict) -> Image.Image:
    """Crop a face from Rekognition bounding box (with padding)."""
    W, H = img.size
    left = int(bbox["Left"] * W)
    top = int(bbox["Top"] * H)
    width = int(bbox["Width"] * W)
    height = int(bbox["Height"] * H)
    pad = int(0.05 * max(width, height))
    left = max(0, left - pad)
    top = max(0, top - pad)
    right = min(W, left + width + 2 * pad)
    bottom = min(H, top + height + 2 * pad)
    return img.crop((left, top, right, bottom))


def image_to_bytes(img: Image.Image) -> bytes:
    buf = BytesIO()
    img.save(buf, format="JPEG")
    return buf.getvalue()


# ========= Global no-cache headers =========
@app.after_request
def add_no_cache_headers(resp):
    resp.headers["Cache-Control"] = "no-store, no-cache, must-revalidate, max-age=0"
    resp.headers["Pragma"] = "no-cache"
    resp.headers["Expires"] = "0"
    return resp


# ========= Routes =========
@app.route('/')
def index():
    return render_template("index.html")


@app.route("/enroll", methods=["POST"])
def enroll_student():
    if "file" not in request.files or "student_id" not in request.form or "name" not in request.form:
        return jsonify({"error": "Image, student_id, and name are required"}), 400

    file = request.files["file"]
    student_id = request.form["student_id"].strip()
    name = request.form["name"].strip()
    image_bytes = file.read()

    # Save reference image locally
    file_path = os.path.join(REF_DIR, f"{student_id}_{name}.jpg")
    with open(file_path, "wb") as f:
        f.write(image_bytes)

    # Index into Rekognition
    response = rekognition.index_faces(
        CollectionId=COLLECTION_ID,
        Image={"Bytes": image_bytes},
        ExternalImageId=student_id,
        DetectionAttributes=["DEFAULT"],
        MaxFaces=5
    )
    save_student(student_id, name)

    return jsonify({
        "message": f"Student {name} enrolled successfully",
        "student_id": student_id,
        "image_saved": file_path,
        "rekognition_response": response
    })


@app.route("/detect", methods=["POST"])
def detect_and_mark():
    """Detect faces, identify students, log attendance, and update concentration stats."""
    if "file" not in request.files:
        return jsonify({"error": "No file"}), 400

    img_bytes = request.files["file"].read()

    # Detect faces
    face_details = rekognition.detect_faces(Image={"Bytes": img_bytes}, Attributes=["ALL"])
    details = face_details.get("FaceDetails", [])

    if not details:
        return jsonify({
            "faces": [],
            "stats": {
                "focused_total": sum(FOCUS_SECONDS.values()),
                "distracted_total": sum(DISTRACTED_SECONDS.values()),
                "distracted_students": []
            }
        })

    base_img = Image.open(BytesIO(img_bytes)).convert("RGB")
    faces_payload = []
    distracted_names_now = []

    for fd in details:
        bbox = fd["BoundingBox"]
        eyes_open = fd.get("EyesOpen", {}).get("Value", False)
        is_focused = bool(eyes_open)

        # Identify student
        try:
            face_crop = pil_crop_from_normalized_bbox(base_img, bbox)
            crop_bytes = image_to_bytes(face_crop)
            match_resp = rekognition.search_faces_by_image(
                CollectionId=COLLECTION_ID,
                Image={"Bytes": crop_bytes},
                MaxFaces=1,
                FaceMatchThreshold=90
            )
            if match_resp.get("FaceMatches"):
                match = match_resp["FaceMatches"][0]
                student_id = match["Face"]["ExternalImageId"]
                similarity = float(match["Similarity"])
                name = get_student_name(student_id)
            else:
                student_id = "unknown"
                similarity = 0.0
                name = "Unknown"
        except Exception:
            student_id = "unknown"
            similarity = 0.0
            name = "Unknown"

        # Log attendance
        log_attendance_once(student_id, name, "Concentrated ✅" if is_focused else "Not Concentrated ❌")

        # Update session counters
        if is_focused:
            FOCUS_SECONDS[student_id] = FOCUS_SECONDS.get(student_id, 0) + POLL_INTERVAL_SEC
        else:
            DISTRACTED_SECONDS[student_id] = DISTRACTED_SECONDS.get(student_id, 0) + POLL_INTERVAL_SEC
            if name != "Unknown":
                distracted_names_now.append(name)

        # Payload for frontend
        faces_payload.append({
            "UserID": student_id,
            "Name": name,
            "Similarity": similarity,
            "Concentration": "Focused" if is_focused else "Distracted",
            "BoundingBox": bbox,
            "Gender": fd.get("Gender", {}).get("Value", "Unknown"),
            "AgeRange": fd.get("AgeRange", {"Low": "?", "High": "?"}),
            "FocusTimeSec": FOCUS_SECONDS.get(student_id, 0),
            "DistractedTimeSec": DISTRACTED_SECONDS.get(student_id, 0),
            "LastSeen": datetime.now().strftime("%H:%M:%S")
        })

    stats_payload = {
        "focused_total": sum(FOCUS_SECONDS.values()),
        "distracted_total": sum(DISTRACTED_SECONDS.values()),
        "distracted_students": sorted(set(distracted_names_now))
    }

    return jsonify({"faces": faces_payload, "stats": stats_payload})


@app.route("/download_attendance", methods=["GET"])
def download_attendance():
    ensure_attendance_file()

    @after_this_request
    def no_cache(r):
        r.headers["Cache-Control"] = "no-store, no-cache, must-revalidate, max-age=0"
        r.headers["Pragma"] = "no-cache"
        r.headers["Expires"] = "0"
        return r

    return send_file(ATTENDANCE_FILE, as_attachment=True)


if __name__ == "__main__":
    # Disable Flask reloader to avoid duplicate state
    app.run(debug=True, use_reloader=False)
